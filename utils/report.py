# ---------------------------------------------------------------------
# report.py — Visualizations and reporting helpers
# (Plotly + ReportLab + OpenPyXL, headless & browser-independent)
# ---------------------------------------------------------------------

from __future__ import annotations

import io
import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import plotly.io as pio
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, Optional, List, Tuple

# ---------------------------------------------------------------------
# Plotly headless image-export configuration (no Chrome dependency)
# ---------------------------------------------------------------------
try:
    # Modern API (Plotly ≥ 5.23)
    pio.defaults.renderers.default = "png"
    pio.defaults.to_image.default_engine = "kaleido"
    pio.defaults.to_image.chromium_args = [
        "--no-sandbox",
        "--disable-gpu",
        "--disable-dev-shm-usage",
    ]
except Exception:
    # Backward compatibility (< 5.23)
    try:
        if hasattr(pio, "kaleido"):
            pio.kaleido.scope.default_format = "png"
            pio.kaleido.scope.chromium_args = [
                "--no-sandbox",
                "--disable-gpu",
                "--disable-dev-shm-usage",
            ]
    except Exception:
        pass


# ---------- Safe Plotly → PNG converter ----------

def _fig_to_png_bytes(fig, scale: float = 2.0) -> bytes:
    """
    Chrome-free fallback exporter.
    Tries Kaleido once; on any failure, inserts a neutral placeholder image.
    """
    import io
    from PIL import Image
    import plotly.io as pio

    try:
        # Try normal Kaleido render (works if deps present)
        return pio.to_image(fig, format="png", scale=scale, engine="kaleido")
    except Exception as e:
        # Fallback: neutral placeholder so report generation never fails
        img = Image.new("RGB", (960, 540), color=(245, 245, 245))
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)
        return buf.getvalue()




# ---------- Plot Builders ----------

def plot_churn_density(df: pd.DataFrame, prob_col: str = "churn_prob", threshold: float = 0.5):
    if prob_col not in df.columns:
        raise KeyError(f"Column '{prob_col}' not found for churn density.")
    fig = px.histogram(
        df, x=prob_col, nbins=40, histnorm="probability density",
        title="Churn Probability — Density with Threshold"
    )
    fig.add_vline(x=threshold, line_width=2, line_dash="dash", annotation_text=f"thr={threshold:.2f}")
    fig.update_layout(xaxis_title="Churn Probability", yaxis_title="Density", bargap=0.02)
    return fig


def plot_feature_importance(imp_df, top_n=20):
    import plotly.express as px
    if imp_df is None or imp_df.empty:
        return None

    df = imp_df.head(top_n).copy()
    df["Importance"] = df["Importance"] / df["Importance"].sum()

    fig = px.bar(
        df,
        x="Importance",
        y="Feature",
        orientation="h",
        title=f"Top {top_n} Feature Importances",
    )
    fig.update_layout(
        yaxis={"categoryorder": "total ascending"},
        xaxis_title="Relative Importance",
        yaxis_title="Feature",
    )
    return fig



def plot_risk_value_quadrant(
    df: pd.DataFrame,
    risk_col: str = "churn_prob",
    value_col: str = "predicted_clv",
    seg_col: str = "segment",
    risk_thr: Optional[float] = None,
    value_thr: Optional[float] = None
):
    for col in (risk_col, value_col, seg_col):
        if col not in df.columns:
            raise KeyError(f"Required column '{col}' missing for quadrant plot.")
    if risk_thr is None:
        risk_thr = float(df[risk_col].median())
    if value_thr is None:
        value_thr = float(df[value_col].median())

    fig = px.scatter(
        df, x=risk_col, y=value_col, color=seg_col,
        hover_data=[c for c in df.columns if c not in [risk_col, value_col]],
        title="Risk vs Value — Quadrant View"
    )
    fig.add_vline(x=risk_thr, line_width=2, line_dash="dash")
    fig.add_hline(y=value_thr, line_width=2, line_dash="dash")
    fig.update_layout(xaxis_title="Churn Probability", yaxis_title="Predicted CLV")
    return fig


def plot_clv_distribution(df: pd.DataFrame, clv_col: str = "predicted_clv", log_scale: bool = False):
    if clv_col not in df.columns:
        raise KeyError(f"Column '{clv_col}' not found for CLV distribution.")
    fig = px.histogram(df, x=clv_col, nbins=40, title="CLV Distribution")
    fig.update_layout(xaxis_title="Predicted CLV", yaxis_title="Count")
    if log_scale:
        fig.update_xaxes(type="log")
    return fig


def plot_retention_curve(df: pd.DataFrame, tenure_col: Optional[str], target_col: Optional[str]) -> Optional[go.Figure]:
    """
    Simple survival-like curve using tenure buckets. Returns None if tenure/target unavailable.
    Assumes target 1=churned, 0=retained.
    """
    if not tenure_col or not target_col:
        return None
    if tenure_col not in df.columns or target_col not in df.columns:
        return None

    temp = df[[tenure_col, target_col]].dropna()
    ten = pd.to_numeric(temp[tenure_col], errors="coerce")
    keep = ~ten.isna()
    temp = temp.loc[keep].copy()
    temp["tenure_num"] = ten[keep].astype(float)
    temp["tenure_m"] = temp["tenure_num"].round().astype(int)

    grp = temp.groupby("tenure_m")[target_col].mean().reset_index()
    grp["retention"] = 1.0 - grp[target_col]
    grp = grp.sort_values("tenure_m")

    if grp.empty:
        return None

    fig = go.Figure()
    fig.add_trace(go.Scatter(x=grp["tenure_m"], y=grp["retention"], mode="lines+markers", name="Retention"))
    fig.update_layout(title="Retention Curve (by Tenure)", xaxis_title="Tenure (months)", yaxis_title="Retention")
    return fig


# ---------- Summary Builders ----------

def build_summary_table(metrics: Dict[str, float], portfolio: Dict[str, float]) -> pd.DataFrame:
    rows = [
        ("AUC", metrics.get("AUC", float("nan"))),
        ("Precision", metrics.get("Precision", float("nan"))),
        ("Recall", metrics.get("Recall", float("nan"))),
        ("F1", metrics.get("F1", float("nan"))),
        ("Total Expected Value", portfolio.get("total_expected_value", float("nan"))),
        ("Value Concentration (Top 10%)", portfolio.get("value_concentration_top_share", float("nan")) * 100.0),
        ("High-Value Retention", portfolio.get("high_value_retention_rate", float("nan")) * 100.0),
        ("Avg Expected Months", portfolio.get("avg_expected_months", float("nan"))),
    ]
    df = pd.DataFrame(rows, columns=["Metric", "Value"])
    # Format percentages for nicer display in Streamlit table if desired — keep raw numeric for Excel.
    return df


# ---------- Recommendation Engine ----------

def generate_recommendations(
    metrics: Dict[str, float],
    portfolio: Dict[str, float],
    risk_threshold: float,
    value_threshold: float
) -> List[str]:
    recs: List[str] = []
    # Model quality gates
    if metrics.get("Recall", 0.0) < 0.70:
        recs.append("Raise recall: add behavioral features (usage, support tickets, payment history). Consider class-weight or focal loss.")
    if metrics.get("Precision", 0.0) < 0.60:
        recs.append("Raise precision: increase decision threshold and apply probability calibration (isotonic).")
    if metrics.get("AUC", 0.0) < 0.75:
        recs.append("Discrimination moderate: optimize XGBoost depth/learning-rate; evaluate monotonic constraints for stability.")

    # Portfolio economics
    conc = portfolio.get("value_concentration_top_share", 0.0)
    hvr = portfolio.get("high_value_retention_rate", 0.0)
    if conc > 0.50:
        recs.append("Value concentrated: launch VIP retention for top 10% (priority support, dedicated account mgmt).")
    if hvr < 0.90:
        recs.append("High-value retention below target: offer targeted save-offers to 'High Risk • High Value' cohort.")

    # Ops & governance
    recs.append("Monitor drift via PSI monthly; retrain when PSI > 0.2 or cohort mix changes.")
    recs.append(f"Prioritize customers above risk>{risk_threshold:.2f} and value>{value_threshold:,.0f} for intervention.")
    return recs


# ---------- Report Builders (Excel + PDF) ----------

def _fig_to_png_bytes(fig: go.Figure, scale: float = 2.0) -> bytes:
    """Convert Plotly figure to PNG bytes via kaleido."""
    return pio.to_image(fig, format="png", scale=scale)


def build_excel_report(
    summary_df: pd.DataFrame,
    top_risk_df: pd.DataFrame,
    top_value_df: pd.DataFrame,
    segments_df: pd.DataFrame,
    full_scores_df: Optional[pd.DataFrame] = None,
) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    for row in dataframe_to_rows(summary_df, index=False, header=True):
        ws.append(row)

    ws2 = wb.create_sheet("Top At-Risk")
    for row in dataframe_to_rows(top_risk_df, index=False, header=True):
        ws2.append(row)

    ws3 = wb.create_sheet("Top Value")
    for row in dataframe_to_rows(top_value_df, index=False, header=True):
        ws3.append(row)

    ws4 = wb.create_sheet("Segments")
    for row in dataframe_to_rows(segments_df, index=False, header=True):
        ws4.append(row)

    if full_scores_df is not None:
        ws5 = wb.create_sheet("All Scores (sample)")
        # limit to avoid huge files
        head = full_scores_df.head(1000)
        for row in dataframe_to_rows(head, index=False, header=True):
            ws5.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_pdf_report(
    title: str,
    subtitle_lines: List[str],
    summary_rows: List[Tuple[str, str]],
    figs: List[Tuple[str, go.Figure]],
    tables: List[Tuple[str, pd.DataFrame]],
    page_orientation: str = "portrait",  # 'portrait' or 'landscape'
) -> io.BytesIO:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas
    from reportlab.platypus import Table, TableStyle

    pagesize = A4 if page_orientation == "portrait" else landscape(A4)
    width, height = pagesize

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=pagesize)

    # Header
    c.setFont("Helvetica-Bold", 16)
    c.drawString(2*cm, height - 2*cm, title)
    c.setFont("Helvetica", 10)
    y = height - 2.7*cm
    for line in subtitle_lines:
        c.drawString(2*cm, y, line)
        y -= 0.5*cm

    # Summary table
    data = [["Metric", "Value"]] + [[k, v] for k, v in summary_rows]
    table = Table(data, colWidths=[7*cm, 8*cm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("TEXTCOLOR", (0,0), (-1,0), colors.black),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,0), 10),
        ("ALIGN", (0,0), (-1,-1), "LEFT"),
        ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
        ("BOTTOMPADDING", (0,0), (-1,0), 6),
    ]))
    w, h = table.wrapOn(c, width - 4*cm, height)
    table.drawOn(c, 2*cm, y - h)
    y = y - h - 0.8*cm

    # Figures — each on its own area; add new page if needed
    for caption, fig in figs:
        png = _fig_to_png_bytes(fig, scale=2.0)
        img_w = width - 4*cm
        img_h = img_w * 0.56  # 16:9-ish
        if y - img_h < 2*cm:
            c.showPage()
            c.setFont("Helvetica", 10)
            y = height - 2*cm
        c.setFont("Helvetica-Bold", 11)
        c.drawString(2*cm, y, caption)
        y -= 0.5*cm
        c.drawImage(io.BytesIO(png), 2*cm, y - img_h, width=img_w, height=img_h, preserveAspectRatio=True, mask='auto')
        y -= img_h + 0.6*cm

    # Tables (truncate to reasonable length)
    for caption, df in tables:
        if y < 6*cm:
            c.showPage()
            c.setFont("Helvetica", 10)
            y = height - 2*cm
        c.setFont("Helvetica-Bold", 11)
        c.drawString(2*cm, y, caption)
        y -= 0.5*cm

        head = df.head(20).copy()
        data = [head.columns.tolist()] + head.astype(str).values.tolist()
        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
            ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
            ("FONTSIZE", (0,0), (-1,0), 9),
            ("FONTSIZE", (0,1), (-1,-1), 8),
            ("ALIGN", (0,0), (-1,-1), "LEFT"),
        ]))
        tw, th = t.wrapOn(c, width - 4*cm, height)
        t.drawOn(c, 2*cm, y - th)
        y -= th + 0.6*cm

    c.showPage()
    c.save()
    buf.seek(0)
    return buf
